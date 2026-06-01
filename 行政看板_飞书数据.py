#!/usr/bin/env python3
"""行政数据看板 — 飞书实时数据脚本
从飞书「行政数据-实时更新」电子表格直接拉取费用数据，
同时从本地 Excel 读取电脑台账和固定资产台账，
合并输出 dashboard_data.json 供前端看板使用。
"""
import json
import os
import subprocess
import sys
from collections import defaultdict
from datetime import datetime

# ============ 配置 ============
LARK_CLI = "/Users/duolaameng/.workbuddy/binaries/node/versions/22.12.0/bin/lark-cli"
SPREADSHEET_TOKEN = "MTUls4SkvhMybJtK9EjcChqonnc"
SHEET_ID = "6d1b5c"
EXCEL_FILE = "/Users/duolaameng/Desktop/文件/作业-小五/资产管理台账字段.xlsx"
OUTPUT_DIR = "/Users/duolaameng/Desktop/文件/作业-小五/看板数据"
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "dashboard_data.json")

os.makedirs(OUTPUT_DIR, exist_ok=True)


def run_lark(*args):
    """运行 lark-cli 命令并返回 JSON"""
    cmd = [LARK_CLI] + list(args) + ["--as", "user"]
    env = os.environ.copy()
    env.pop("NODE_OPTIONS", None)  # 避免 node options 冲突
    result = subprocess.run(cmd, capture_output=True, text=True, env=env, timeout=30)
    if result.returncode != 0:
        print(f"  ⚠️ lark-cli 错误: {result.stderr[:200]}")
        return None
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        print(f"  ⚠️ JSON 解析失败: {result.stdout[:200]}")
        return None


def fetch_feishu_expenses():
    """从飞书电子表格拉取费用数据"""
    print("📡 从飞书拉取费用数据...")

    # 先获取表格信息确定行数
    info = run_lark("sheets", "+info", "--spreadsheet-token", SPREADSHEET_TOKEN)
    if not info or not info.get("ok"):
        print("  ❌ 无法获取飞书表格信息")
        return []

    row_count = info["data"]["sheets"]["sheets"][0]["grid_properties"]["row_count"]

    # 读取全部数据（从第2行开始）
    range_str = f"A2:T{row_count}"
    data = run_lark(
        "sheets", "+read",
        "--spreadsheet-token", SPREADSHEET_TOKEN,
        "--sheet-id", SHEET_ID,
        "--range", range_str
    )

    if not data or not data.get("ok"):
        print("  ❌ 无法读取飞书表格数据")
        return []

    values = data["data"]["valueRange"].get("values", [])
    records = []

    for row in values:
        if not row or row[0] is None:
            continue
        # 飞书表格列映射:
        # A=序号 B=流程编号 C=付款年 D=付款月 E=付款日期 F=不含税金额
        # G=申请员工 H=部门 I=付款金额 J=分公司 K=费用项目 L=付款用途
        # M=流程状态 N=归属主体 O=收款公司 P=是否取得发票 Q=发票号
        # R=发票类型 S=发票是否移交财务 T=移交方式
        try:
            # 安全的取列值
            def col(i, default=None):
                return row[i] if i < len(row) else default

            # 付款年份：优先从C列取，如果是Excel序列号则从E列日期推算
            raw_year = col(2)
            raw_date = col(4)
            year = 0
            month = 0
            date_str = ""

            # 尝试解析日期
            if raw_date is not None:
                try:
                    date_num = float(raw_date)
                    # Excel序列号转日期 (1899-12-30基准)
                    if 40000 < date_num < 60000:
                        from datetime import timedelta
                        base = datetime(1899, 12, 30)
                        real_date = base + timedelta(days=int(date_num))
                        year = real_date.year
                        month = real_date.month
                        date_str = real_date.strftime("%Y-%m-%d")
                    else:
                        date_str = str(raw_date)[:10]
                except (ValueError, OverflowError):
                    date_str = str(raw_date)[:10]

            # 如果日期解析失败，尝试从C/D列获取年/月
            if year == 0 and raw_year is not None:
                try:
                    yr = int(float(raw_year))
                    if 2020 <= yr <= 2030:
                        year = yr
                except (ValueError, OverflowError):
                    pass

            if month == 0:
                raw_month = col(3)
                if raw_month is not None:
                    try:
                        m = int(float(raw_month))
                        if 1 <= m <= 12:
                            month = m
                    except (ValueError, OverflowError):
                        pass

            # 金额：I列（付款金额）
            amount = 0
            raw_amount = col(8)
            if raw_amount is not None:
                try:
                    amount = float(raw_amount)
                except (ValueError, TypeError):
                    pass

            rec = {
                "seq": col(0),
                "flow_id": str(col(1) or "").strip(),
                "year": year,
                "month": month,
                "date": date_str,
                "amount_ex_tax": float(col(5)) if col(5) else 0,
                "applicant": str(col(6) or "").strip(),
                "dept": str(col(7) or "").strip(),
                "amount": amount,
                "city": str(col(9) or "").strip(),
                "category": str(col(10) or "").strip(),
                "purpose": str(col(11) or "")[:80].strip(),
                "status": str(col(12) or "").strip(),
                "entity": str(col(13) or "").strip(),
                "supplier": str(col(14) or "").strip(),
                "invoice": str(col(15) or "").strip(),
                "invoice_no": str(col(16) or "").strip(),
                "invoice_type": str(col(17) or "").strip(),
                "invoice_transfer": str(col(18) or "").strip(),
                "transfer_method": str(col(19) or "").strip(),
            }

            if rec["year"] >= 2023:
                records.append(rec)
        except (TypeError, ValueError, IndexError) as e:
            print(f"  ⚠️ 跳过一行数据: {e}")
            continue

    print(f"  ✅ 从飞书读取 {len(records)} 条费用记录")
    return records


def process_computers_from_excel():
    """从本地 Excel 读取电脑台账"""
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["电脑管理台账"]
    computers = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            comp = {
                "seq": row[0],
                "asset_no": str(row[1] or ""),
                "category": str(row[2] or ""),
                "name": str(row[3] or ""),
                "spec": str(row[4] or ""),
                "qty": int(row[5]) if row[5] else 1,
                "unit": str(row[6] or ""),
                "asset_type": str(row[7] or ""),
                "supplier": str(row[8] or ""),
                "user": str(row[9] or ""),
                "city": str(row[10] or ""),
                "dept": str(row[11] or ""),
                "sub_dept": str(row[12] or ""),
                "sub_sub_dept": str(row[13] or ""),
                "emp_no": str(row[14] or ""),
                "status": str(row[15] or ""),
                "value": float(row[16]) if row[16] else 0,
                "start_date": str(row[17] or "")[:10],
                "end_date": str(row[18] or "")[:10],
                "remark": str(row[19] or "")[:30],
            }
            computers.append(comp)
        except (TypeError, ValueError):
            continue

    return computers


def process_fixed_assets_from_excel():
    """从本地 Excel 读取固定资产台账"""
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
        import openpyxl

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb["固定资产台账"]
    assets = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            a = {
                "seq": row[0],
                "asset_no": str(row[1] or ""),
                "category": str(row[2] or ""),
                "name": str(row[3] or ""),
                "spec": str(row[4] or ""),
                "asset_type": str(row[5] or ""),
                "qty": int(row[6]) if row[6] else 1,
                "unit": str(row[7] or ""),
                "value": float(row[8]) if row[8] else 0,
                "start_date": str(row[9] or "")[:10],
                "end_date": str(row[10] or "")[:10],
                "supplier": str(row[11] or ""),
                "city": str(row[12] or ""),
                "location": str(row[13] or "")[:30],
                "dept": str(row[14] or ""),
                "person": str(row[15] or ""),
                "status": str(row[16] or ""),
                "remark": str(row[17] or "")[:30],
            }
            assets.append(a)
        except (TypeError, ValueError):
            continue

    return assets


def _aggregate_for_year(records, year):
    """为指定年份聚合数据"""
    yr_records = [r for r in records if r["year"] == year]
    if not yr_records:
        return None

    monthly = defaultdict(float)
    monthly_count = defaultdict(int)
    by_category = defaultdict(float)
    by_category_count = defaultdict(int)
    by_city = defaultdict(float)
    by_city_count = defaultdict(int)
    daily = defaultdict(float)
    daily_count = defaultdict(int)

    for r in yr_records:
        key = f"{r['month']:02d}"
        monthly[key] += r["amount"]
        monthly_count[key] += 1

        cat_short = r["category"].replace("日常行政费用/", "").replace("保证金/", "")
        if "/" in cat_short:
            cat_short = cat_short.split("/")[0]
        by_category[cat_short] += r["amount"]
        by_category_count[cat_short] += 1

        city_clean = r["city"].replace("市", "").strip()
        by_city[city_clean] += r["amount"]
        by_city_count[city_clean] += 1

        day_key = r["date"][:10] if r["date"] else ""
        if day_key:
            daily[day_key] += r["amount"]
            daily_count[day_key] += 1

    monthly_sorted = sorted(monthly.items(), key=lambda x: x[0])
    by_city_sorted = sorted(by_city.items(), key=lambda x: -x[1])
    by_category_sorted = sorted(by_category.items(), key=lambda x: -x[1])
    daily_sorted = sorted(daily.items(), key=lambda x: x[0])

    total_amount = sum(r["amount"] for r in yr_records)
    total_count = len(yr_records)
    filled_months = len(monthly)
    avg_monthly = round(total_amount / filled_months, 2) if filled_months else 0

    today = datetime.now().strftime("%Y-%m-%d")
    today_amount = daily.get(today, 0)

    records_sorted = sorted(yr_records, key=lambda x: x["date"], reverse=True)
    latest = [
        {
            "date": r["date"],
            "supplier": r["supplier"][:20],
            "city": r["city"],
            "category": r["category"].replace("日常行政费用/", ""),
            "amount": round(r["amount"], 2),
            "purpose": r["purpose"][:50],
            "status": r["status"],
        }
        for r in records_sorted[:20]
    ]

    return {
        "totalAmount": round(total_amount, 2),
        "totalCount": total_count,
        "avgMonthly": avg_monthly,
        "filledMonths": filled_months,
        "thisMonthTotal": round(today_amount, 2),
        "todayAmount": round(today_amount, 2),
        "monthlyTrend": [
            {"label": f"{year}-{k}", "amount": round(v, 2), "count": monthly_count[k]}
            for k, v in monthly_sorted
        ],
        "categoryDistribution": [
            {"name": k, "amount": round(v, 2), "count": by_category_count[k]}
            for k, v in by_category_sorted[:15]
        ],
        "cityDistribution": [
            {"name": k if k else "未知", "amount": round(v, 2), "count": by_city_count[k]}
            for k, v in by_city_sorted[:15]
        ],
        "daily": [
            {"date": k, "amount": round(v, 2), "count": daily_count[k]}
            for k, v in daily_sorted[-90:]
        ],
        "latestExpenses": latest,
    }


def detect_duplicates(records):
    """检测疑似重复费用记录。

    分级规则：
    - 高危(high)：同供应商 + 同金额(分) + 同月 + 不同流程编号
    - 中危(medium)：同城市 + 同类别 + 同金额 + 同月 + 不同流程编号
    - 低危(low)：同供应商 + 同金额 + 同季度

    返回: { "summary": {...}, "byYear": {...}, "groups": [...] }
    """
    if not records:
        return {"summary": {"totalGroups": 0, "totalRecords": 0, "totalAmount": 0, "highGroups": 0, "mediumGroups": 0, "lowGroups": 0}, "byYear": {}, "groups": []}

    # 过滤掉金额为 0 的记录
    valid = [r for r in records if r["amount"] > 0]

    # 高危：同供应商 + 同金额 + 同月
    high_map = defaultdict(list)
    for r in valid:
        key = (r["supplier"], round(r["amount"], 2), r["year"], r["month"])
        high_map[key].append(r)

    high_groups = []
    for key, recs in high_map.items():
        if len(recs) < 2:
            continue
        flow_ids = set(r["flow_id"] for r in recs)
        if len(flow_ids) < 2:
            continue
        high_groups.append(recs)

    # 中危：同城市 + 同类别(前两级) + 同金额 + 同月（排除已在高危中的）
    high_flow_ids = set()
    for g in high_groups:
        for r in g:
            high_flow_ids.add(r["flow_id"])

    medium_map = defaultdict(list)
    for r in valid:
        if r["flow_id"] in high_flow_ids:
            continue
        cat_short = r["category"].split("/")[0] if "/" in r["category"] else r["category"]
        key = (r["city"], cat_short, round(r["amount"], 2), r["year"], r["month"])
        medium_map[key].append(r)

    medium_groups = []
    for key, recs in medium_map.items():
        if len(recs) < 2:
            continue
        flow_ids = set(r["flow_id"] for r in recs)
        if len(flow_ids) < 2:
            continue
        medium_groups.append(recs)

    # 合并并按记录数排序
    all_groups = []
    for g in high_groups:
        all_groups.append({"level": "high", "records": g})
    for g in medium_groups:
        all_groups.append({"level": "medium", "records": g})

    all_groups.sort(key=lambda x: -len(x["records"]))

    # 构建输出（TOP 50 组）
    groups_out = []
    for g in all_groups[:50]:
        recs = g["records"]
        supplier = recs[0]["supplier"][:30]
        amt = round(recs[0]["amount"], 2)
        yr = recs[0]["year"]
        mo = recs[0]["month"]
        total_amt = round(sum(r["amount"] for r in recs), 2)
        items = []
        for r in sorted(recs, key=lambda x: x["date"]):
            items.append({
                "date": r["date"],
                "city": r["city"],
                "category": r["category"].replace("日常行政费用/", ""),
                "amount": round(r["amount"], 2),
                "purpose": r["purpose"][:50],
                "supplier": r["supplier"][:20],
            })
        groups_out.append({
            "level": g["level"],
            "supplier": supplier,
            "amount": amt,
            "year": yr,
            "month": mo,
            "count": len(recs),
            "totalAmount": total_amt,
            "items": items,
        })

    # 汇总统计
    all_recs = set()
    total_amt = 0
    high_count = 0
    medium_count = 0
    for g in all_groups:
        for r in g["records"]:
            all_recs.add(r["flow_id"])
            total_amt += r["amount"]
        if g["level"] == "high":
            high_count += 1
        else:
            medium_count += 1

    # 按年份统计
    by_year = defaultdict(lambda: {"groups": 0, "records": 0, "amount": 0})
    for g in all_groups:
        yr = str(g["records"][0]["year"])
        by_year[yr]["groups"] += 1
        by_year[yr]["records"] += len(g["records"])
        by_year[yr]["amount"] += round(sum(r["amount"] for r in g["records"]), 2)

    return {
        "summary": {
            "totalGroups": len(all_groups),
            "totalRecords": len(all_recs),
            "totalAmount": round(total_amt, 2),
            "highGroups": high_count,
            "mediumGroups": medium_count,
            "lowGroups": 0,
        },
        "byYear": {yr: by_year[yr] for yr in sorted(by_year)},
        "groups": groups_out,
    }


def aggregate_expenses(records):
    """聚合费用数据 — 全量 + 按年份"""
    all_years = sorted(set(r["year"] for r in records if r["year"] >= 2023))

    # 重复检测
    dup_check = detect_duplicates(records)

    # 按年份分别聚合
    by_year = {}
    for yr in all_years:
        yr_data = _aggregate_for_year(records, yr)
        if yr_data:
            by_year[str(yr)] = yr_data

    # 全量聚合（保留向后兼容）
    monthly = defaultdict(float)
    monthly_count = defaultdict(int)
    by_category = defaultdict(float)
    by_category_count = defaultdict(int)
    by_city = defaultdict(float)
    by_city_count = defaultdict(int)
    daily_2026 = defaultdict(float)
    daily_2026_count = defaultdict(int)

    for r in records:
        key = f"{r['year']}-{r['month']:02d}"
        monthly[key] += r["amount"]
        monthly_count[key] += 1
        cat_short = r["category"].replace("日常行政费用/", "").replace("保证金/", "")
        if "/" in cat_short:
            cat_short = cat_short.split("/")[0]
        by_category[cat_short] += r["amount"]
        by_category_count[cat_short] += 1
        city_clean = r["city"].replace("市", "").strip()
        by_city[city_clean] += r["amount"]
        by_city_count[city_clean] += 1
        if r["year"] == 2026:
            day_key = r["date"][:10] if r["date"] else ""
            if day_key:
                daily_2026[day_key] += r["amount"]
                daily_2026_count[day_key] += 1

    monthly_sorted = sorted(monthly.items(), key=lambda x: x[0])
    by_city_sorted = sorted(by_city.items(), key=lambda x: -x[1])
    by_category_sorted = sorted(by_category.items(), key=lambda x: -x[1])
    daily_2026_sorted = sorted(daily_2026.items(), key=lambda x: x[0])

    total_amount = sum(r["amount"] for r in records)
    total_count = len(records)

    amounts_2025 = [r["amount"] for r in records if r["year"] == 2025]
    avg_monthly_2025 = sum(amounts_2025) / 12 if amounts_2025 else 0

    today = datetime.now().strftime("%Y-%m-%d")
    today_amount = daily_2026.get(today, 0)

    records_sorted = sorted(records, key=lambda x: x["date"], reverse=True)
    latest = [
        {
            "date": r["date"],
            "supplier": r["supplier"][:20],
            "city": r["city"],
            "category": r["category"].replace("日常行政费用/", ""),
            "amount": round(r["amount"], 2),
            "purpose": r["purpose"][:50],
            "status": r["status"],
        }
        for r in records_sorted[:20]
        if r["year"] >= 2025
    ]

    return {
        "totalAmount": round(total_amount, 2),
        "totalCount": total_count,
        "avgMonthly": round(avg_monthly_2025, 2),
        "thisMonthTotal": 0,
        "thisMonthCount": 0,
        "todayAmount": round(today_amount, 2),
        "monthlyTrend": [
            {"label": k, "amount": round(v, 2), "count": monthly_count[k]}
            for k, v in monthly_sorted
        ],
        "categoryDistribution": [
            {"name": k, "amount": round(v, 2), "count": by_category_count[k]}
            for k, v in by_category_sorted[:15]
        ],
        "cityDistribution": [
            {"name": k if k else "未知", "amount": round(v, 2), "count": by_city_count[k]}
            for k, v in by_city_sorted[:15]
        ],
        "daily2026": [
            {"date": k, "amount": round(v, 2), "count": daily_2026_count[k]}
            for k, v in daily_2026_sorted[-90:]
        ],
        "latestExpenses": latest,
        "availableYears": all_years,
        "byYear": by_year,
        "dupCheck": dup_check,
    }


def aggregate_computers(computers):
    """聚合电脑数据"""
    by_status = defaultdict(int)
    by_city = defaultdict(int)

    for c in computers:
        s = c["status"] if c["status"] else "未知"
        city = c["city"].replace("市", "").strip() if c["city"] else "未知"
        by_status[s] += c["qty"]
        by_city[city] += c["qty"]

    total = sum(by_status.values())
    in_use = by_status.get("在用", 0)
    idle = by_status.get("闲置", 0)
    retired = by_status.get("退租", 0) + by_status.get("回收", 0)

    return {
        "total": total,
        "inUse": in_use,
        "idle": idle,
        "retired": retired,
        "lost": by_status.get("丢失", 0),
        "byStatus": [
            {"name": k, "count": v}
            for k, v in sorted(by_status.items(), key=lambda x: -x[1])
        ],
        "byCity": [
            {"name": k, "count": v}
            for k, v in sorted(by_city.items(), key=lambda x: -x[1])[:15]
        ],
    }


def main():
    print("📊 行政数据看板 — 飞书实时数据提取")
    print(f"  时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # 1. 从飞书拉取费用数据
    expenses = fetch_feishu_expenses()

    # 2. 从本地 Excel 读取电脑和固定资产
    print("📋 从本地 Excel 读取资产数据...")
    computers = process_computers_from_excel()
    print(f"  ✅ {len(computers)} 条电脑记录")
    assets = process_fixed_assets_from_excel()
    print(f"  ✅ {len(assets)} 条固定资产记录")

    # 3. 聚合
    print("📊 聚合数据...")
    result = {
        "expenses": aggregate_expenses(expenses),
        "computers": aggregate_computers(computers),
        "lastUpdated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source": "feishu",  # 标记数据来源
    }

    # 4. 输出 JSON
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"\n✅ 数据已保存至: {OUTPUT_FILE}")
    print(f'   费用总额: ¥{result["expenses"]["totalAmount"]:,.2f}')
    print(f'   费用记录: {result["expenses"]["totalCount"]} 条')
    print(f'   电脑总数: {result["computers"]["total"]} 台')
    print(f'   数据来源: 飞书实时')
    print(f'   更新时间: {result["lastUpdated"]}')

    return result


if __name__ == "__main__":
    main()
